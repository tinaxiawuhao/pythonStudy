from openpyxl import load_workbook
from openpyxl import Workbook

# 读取xlsx
wb = load_workbook("template.xlsx")
sheet = wb.get_sheet_by_name("Sheet3")
print(sheet["C"])  # (<Cell Sheet3.C1>, <Cell Sheet3.C2>, <Cell Sheet3.C3>, <Cell Sheet3.C4>, <Cell Sheet3.C5>, <Cell Sheet3.C6>, <Cell Sheet3.C7>, <Cell Sheet3.C8>, <Cell Sheet3.C9>, <Cell Sheet3.C10>)      <-第C列
print(sheet["4"])  # (<Cell Sheet3.A4>, <Cell Sheet3.B4>, <Cell Sheet3.C4>, <Cell Sheet3.D4>, <Cell Sheet3.E4>)     <-第4行
print(sheet["C4"].value)  # c4     <-第C4格的值
print(sheet.max_row)  # 10     <-最大行数
print(sheet.max_column)  # 5     <-最大列数
for i in sheet["C"]:
    print(i.value, end=" ")  # c1 c2 c3 c4 c5 c6 c7 c8 c9 c10     <-C列中的所有值

# 写入xlsx
wb = Workbook()
sheet = wb.active
sheet.title = "20210318"

sheet['C3'] = 'Hello world!'
for i in range(10):
    sheet["A%d" % (i + 1)].value = i + 1
# 我们还可以进行花式操作，比如写写公式：
sheet["E1"].value = "=SUM(A:A)"
# 最后记得保存
wb.save('保存一个新的excel.xlsx')
